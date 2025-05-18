from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.messages import constants
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import CreateView
from django.views.generic import ListView
from django.views.generic import UpdateView
from django.views.generic import DeleteView
from django.views.generic import DetailView
from django.http import HttpResponse
from django.db.models import ProtectedError
from ipplan.models import *
from ipplan.forms import *
from src.ipplan.func import *
from ipaddress import ip_network, IPv4Address
import openpyxl
import csv

# Create your views here.


def is_xss_validate(list_string):
    regex = "<([A-Za-z_{}()/]+(|=)*)+>(.*<[A-Za-z/>]+)*"
    for string in list_string:
        result = re.search(regex, string)
        if result:
            return False
    return True


class RegionCreateView(CreateView):
    model = Region
    form_class = RegionForm
    template_name = "create_region.html"
    success_url = "/ipplan/create-region"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = self.request.user
        return super().form_valid(form)


class LocationCreateView(CreateView):
    model = Location
    form_class = LocationForm
    template_name = "create_location.html"
    success_url = "/ipplan/create-location"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = self.request.user
        return super().form_valid(form)


class SubnetCreateView(CreateView):
    model = Subnet
    form_class = SubnetForm
    template_name = "create_subnet.html"
    success_url = "/ipplan/list-subnet"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = self.request.user
        return super().form_valid(form)


class SubnetListView(ListView):
    model = Subnet
    context_object_name = "subnets"
    template_name = "list_subnet.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


@login_required()
def request_ip_form(request):
    if request.method == "POST":
        form = RequestIpAddressForm(request.POST)
        if form.is_valid():
            subnet = form.cleaned_data.get("subnet")
            ips = request.POST.getlist("ip")
            description = form.cleaned_data.get("description")
            for ip in ips:
                IpAddressModel.objects.update_or_create(
                    ip=ip,
                    defaults={
                        "subnet": Subnet.objects.get(subnet=subnet),
                        "inused": True,
                        "description": description,
                        "user_created": str(request.user),
                    },
                )
            return render(
                request,
                template_name="request_ip_result.html",
                context={
                    "ips": ips,
                    "subnet": subnet,
                    "description": description,
                    "user_created": str(request.user),
                },
            )
        else:
            return render(
                request, template_name="request_ip.html", context={"form": form}
            )
    else:
        form = RequestIpAddressForm()
    return render(request, template_name="request_ip.html", context={"form": form})


@login_required()
def dashboard(request):
    db_title = {
        "db_01": "Count location by region",
        "db_02": "Count subnet group by location",
        "db_03": "Top 20 subnets with highest IP usage (%)"
    }
    return render(request, template_name="dashboard.html", context=db_title)


@login_required()
def list_ip(request, pk):
    obj = Subnet.objects.get(id=pk)
    subnet = obj.subnet
    ips = IpAddressModel.objects.filter(subnet__subnet=subnet)
    return render(
        request, template_name="list_ip.html", context={"ips": ips, "subnet": subnet}
    )


class IpAddressModelDeleteView(DeleteView):
    model = IpAddressModel
    template_name = "list_ip.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return IpAddressModel.objects.get(pk=self.kwargs["id"])

    def get_success_url(self):
        pk = self.kwargs["pk"]
        pk1 = self.kwargs["id"]
        return reverse("list-ip", kwargs={"pk": pk})


class IpAddressModelUpdateView(UpdateView):
    model = IpAddressModel
    form_class = IpAddressModelUpdatelForm
    template_name = "update_ip.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return IpAddressModel.objects.get(pk=self.kwargs["id"])

    def get_success_url(self):
        pk = self.kwargs["pk"]
        return reverse("list-ip", kwargs={"pk": pk})


class SubnetUpdateView(UpdateView):
    model = Subnet
    form_class = SubnetForm
    template_name = "update_subnet.html"
    success_url = "/ipplan/list-subnet"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class IpAddressModelDetailView(DetailView):
    model = IpAddressModel
    form_class = IpAddressModelForm
    template_name = "detail_ip.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return IpAddressModel.objects.get(pk=self.kwargs["id"])


class SubnetDeleteView(DeleteView):
    model = Subnet
    template_name = "list-subnet.html"
    success_url = "/ipplan/list-subnet"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)



@login_required()
def create_multiple_subnet(request):
    try:
        if request.method == "POST" and request.FILES.get("upload-file"):
            uploaded_file = request.FILES["upload-file"]
            if str(uploaded_file).lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(uploaded_file)
                worksheet = wb["subnets"]
                for item in worksheet.iter_rows(min_row=2):
                    item = ["" if i.value is None else i.value for i in item]
                    region = item[0]
                    location = item[1]
                    group = item[2]
                    group_subnet = item[3]
                    subnet = item[4]
                    vlan = item[5]
                    name = item[6]
                    description = item[7]
                    user_created = str(request.user)
                    if check_create_multiple_subnet(item):
                        obj_count_01 = Region.objects.filter(
                            region=region).count()
                        obj_count_02 = Location.objects.filter(
                            location=location).count()
                        obj_count_03 = SubnetGroup.objects.filter(
                            group=group).count()
                        obj_count_04 = Subnet.objects.filter(
                            subnet=subnet).count()
                        if obj_count_01 == 0:
                            model = Region(
                                region=region, description=region, user_created=user_created
                            )
                            model.save()
                        if obj_count_02 == 0:
                            obj_01 = Region.objects.get(region=region)
                            model = Location(
                                location=location,
                                region=obj_01,
                                description=region,
                                user_created=user_created,
                            )
                            model.save()
                        if obj_count_03 == 0:
                            obj_02 = Location.objects.get(location=location)
                            model = SubnetGroup(
                                group=group,
                                location=obj_02,
                                group_subnet=group_subnet,
                                description=group,
                                user_created=user_created,
                            )
                            model.save()
                        if obj_count_04 == 0:
                            obj_03 = SubnetGroup.objects.get(group=group)
                            model = Subnet(
                                subnet=subnet,
                                group=obj_03,
                                name=name,
                                vlan=vlan,
                                description=description,
                                user_created=user_created,
                            )
                            model.save()
                        else:
                            model = Subnet.objects.get(subnet=subnet)
                            model.group = SubnetGroup.objects.get(group=group)
                            model.name = name
                            model.vlan = vlan
                            model.description = description
                            model.user_created = user_created
                            model.save()
                messages.add_message(
                    request, constants.SUCCESS, "Import subnet success"
                )
            else:
                messages.add_message(
                    request, constants.ERROR, f"Only support file type *.xlsx")
    except Exception as error:
        messages.add_message(request, constants.ERROR,
                             f"An error occurred: {error}")
    return render(request, template_name="create_multiple_subnet.html")


@login_required()
def request_multiple_ip(request):
    try:
        if request.method == "POST" and request.FILES.get("upload-file"):
            uploaded_file = request.FILES["upload-file"]
            if uploaded_file.lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(uploaded_file)
                worksheet = wb["list_ip"]
                list_subnet = Subnet.objects.all().values_list("subnet", flat=True)
                for item in worksheet.iter_rows(min_row=2):
                    item = ["" if i.value is None else i.value for i in item]
                    description = item[0]
                    ip = item[1]
                    subnet = [
                        subnet
                        for subnet in list_subnet
                        if is_ip(ip) is True
                        and is_subnet(subnet) is True
                        and IPv4Address(ip) in ip_network(subnet)
                    ]
                    if subnet:
                        subnet_obj = Subnet.objects.get(subnet=subnet[-1])
                        IpAddressModel.objects.update_or_create(
                            ip=ip,
                            defaults={
                                "subnet": subnet_obj,
                                "inused": True,
                                "description": description,
                                "user_created": str(request.user),
                            },
                        )
            else:
                messages.add_message(
                    request, constants.ERROR, f"Only support file type *.xlsx")
            messages.add_message(
                request, constants.SUCCESS, "Import IP success")
    except Exception as error:
        messages.add_message(request, constants.ERROR,
                             f"An error occurred: {error}")
    return render(request, template_name="request_multiple_ip.html")


@login_required()
def list_subnet_tree(request):
    context = {}
    if request.method == "GET":
        regions = Region.objects.all()
        locations = Location.objects.all()
        groups = SubnetGroup.objects.all()
        subnets = Subnet.objects.all()
        context = {
            "regions": regions,
            "locations": locations,
            "groups": groups,
            "subnets": subnets,
        }
    return render(request, "list_subnet_tree.html", context=context)


class SubnetGroupCreateView(CreateView):
    model = SubnetGroup
    form_class = SubnetGroupForm
    template_name = "create_subnet_group.html"
    success_url = "/ipplan/create-subnet-group"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)


class SubnetGroupUpdateView(UpdateView):
    model = SubnetGroup
    form_class = SubnetGroupForm
    template_name = "update_subnet_group.html"
    success_url = "/ipplan/list-subnet-group"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)


class SubnetGroupDeleteView(DeleteView):
    model = SubnetGroup
    template_name = "list_subnet_group.html"
    success_url = "/ipplan/list-subnet-group"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class SubnetGroupListView(ListView):
    model = SubnetGroup
    context_object_name = "objects"
    template_name = "list_subnet_group.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


##
class RegionUpdateView(UpdateView):
    model = Region
    form_class = RegionForm
    template_name = "update_region.html"
    success_url = "/ipplan/list-region"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)


class RegionDeleteView(DeleteView):
    model = Region
    template_name = "list_region.html"
    success_url = "/ipplan/list-region"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class RegionListView(ListView):
    model = Region
    context_object_name = "objects"
    template_name = "list_region.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class LocationUpdateView(UpdateView):
    model = Location
    form_class = LocationForm
    template_name = "update_location.html"
    success_url = "/ipplan/list-location"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)


class LocationDeleteView(DeleteView):
    model = Location
    template_name = "list_location.html"
    success_url = "/ipplan/list-location"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class LocationListView(ListView):
    model = Location
    context_object_name = "objects"
    template_name = "list_location.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


@login_required()
def ipplan_export(request):
    form = IPPlanExportForm
    data = {"form": form}
    try:
        if request.method == "POST":
            form = IPPlanExportForm(request.POST)
            if form.is_valid():
                select_id = form.data["database_table"]
                if select_id == "1":
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="subnets.csv"'
                    queryset = Subnet.objects.all()
                    writer = csv.writer(response)
                    writer.writerow([
                        'region',
                        'location',
                        'group',
                        'group_subnet',
                        'subnet',
                        'name',
                        'time_created',
                        'user_created'
                    ])
                    for item in queryset:
                        writer.writerow({
                            item.group.location.region,
                            item.group.location,
                            item.group.group,
                            item.group.group_subnet,
                            item.subnet,
                            item.name,
                            item.time_created,
                            item.user_created
                        })
                    return response
                else:
                    messages.add_message(
                        request, constants.ERROR, 'Database selected is not valid')
    except Exception as error:
        messages.add_message(request, constants.ERROR, error)
    return render(request, template_name="ipplan_export.html", context=data)
