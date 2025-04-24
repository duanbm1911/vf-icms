from django.db.models import ProtectedError
from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.messages import constants
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import CreateView
from django.views.generic import ListView
from django.views.generic import UpdateView
from django.views.generic import DeleteView
from django.views.generic import DetailView
from django.urls import reverse_lazy
from django.urls import reverse
from inventory.models import *
from inventory.forms import *
import csv
import datetime
import openpyxl
import re


MESSAGE_TAGS = {
    messages.ERROR: "danger"
}


def is_xss_validate(list_string):
    regex = "<([A-Za-z_{}()/]+(|=)*)+>(.*<[A-Za-z/>]+)*"
    for string in list_string:
        result = re.search(regex, string)
        if result:
            return False
    return True


class DeviceCreateView(CreateView):
    model = Device
    form_class = DeviceForm
    template_name = "create_device.html"
    success_url = reverse_lazy("list_device")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device"
        return context


class DeviceProvinceCreateView(CreateView):
    model = DeviceProvince
    form_class = DeviceProvinceForm
    template_name = "create_device_province.html"
    success_url = reverse_lazy("list_device_province")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device province"
        return context


class DeviceProvinceUpdateView(UpdateView):
    model = DeviceProvince
    form_class = DeviceProvinceForm
    template_name = "update_device_province.html"
    success_url = reverse_lazy("list_device_province")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device province"
        return context


class DeviceProvinceDeleteView(DeleteView):
    model = DeviceProvince
    template_name = "list_device_province.html"
    success_url = reverse_lazy("list_device_province")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceProvinceListView(ListView):
    model = DeviceProvince
    context_object_name = "objects"
    template_name = "list_device_province.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device provinces"
        return context


class DeviceTypeCreateView(CreateView):
    model = DeviceType
    form_class = DeviceTypeForm
    template_name = "create_device_type.html"
    success_url = reverse_lazy("list_device_type")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device type"
        return context


class DeviceTypeListView(ListView):
    model = DeviceType
    context_object_name = "objects"
    template_name = "list_device_type.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device types"
        return context


class DeviceTypeUpdateView(UpdateView):
    model = DeviceType
    form_class = DeviceTypeForm
    template_name = "update_device_type.html"
    success_url = reverse_lazy("list_device_type")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device type"
        return context


class DeviceTypeDeleteView(DeleteView):
    model = DeviceType
    template_name = "list_device_type.html"
    success_url = reverse_lazy("list_device_type")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceCategoryCreateView(CreateView):
    model = DeviceCategory
    form_class = DeviceCategoryForm
    template_name = "create_device_category.html"
    success_url = reverse_lazy("list_device_category")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device category"
        return context


class DeviceCategoryListView(ListView):
    model = DeviceCategory
    context_object_name = "objects"
    template_name = "list_device_category.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device categories"
        return context


class DeviceCategoryUpdateView(UpdateView):
    model = DeviceCategory
    form_class = DeviceCategoryForm
    template_name = "update_device_category.html"
    success_url = reverse_lazy("list_device_category")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device category"
        return context


class DeviceCategoryDeleteView(DeleteView):
    model = DeviceCategory
    template_name = "list_device_category.html"
    success_url = reverse_lazy("list_device_category")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceVendorCreateView(CreateView):
    model = DeviceVendor
    form_class = DeviceVendorForm
    template_name = "create_device_vendor.html"
    success_url = reverse_lazy("list_device_vendor")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device vendor"
        return context


class DeviceVendorListView(ListView):
    model = DeviceVendor
    context_object_name = "objects"
    template_name = "list_device_vendor.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device vendors"
        return context


class DeviceVendorUpdateView(UpdateView):
    model = DeviceVendor
    form_class = DeviceVendorForm
    template_name = "update_device_vendor.html"
    success_url = reverse_lazy("list_device_vendor")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device vendor"
        return context


class DeviceVendorDeleteView(DeleteView):
    model = DeviceVendor
    template_name = "list_device_vendor.html"
    success_url = reverse_lazy("list_device_vendor")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceListView(ListView):
    model = Device
    context_object_name = "devices"
    template_name = "list_device.html"

    def get_queryset(self):
        queryset = super().get_queryset()
        form = DeviceSearchForm(self.request.GET)
        if not form.is_valid() or not any(form.cleaned_data.values()):
            return Device.objects.none()
        else:
            if form.cleaned_data['device_os']:
                queryset = queryset.filter(
                    device_os__in=form.cleaned_data['device_os'])
            if form.cleaned_data['device_category']:
                queryset = queryset.filter(
                    device_category__in=form.cleaned_data['device_category'])
            if form.cleaned_data['device_group']:
                queryset = queryset.filter(
                    device_group__in=form.cleaned_data['device_group'])
            if form.cleaned_data['device_branch']:
                queryset = queryset.filter(
                    device_branch__in=form.cleaned_data['device_branch'])
            if form.cleaned_data['device_type']:
                queryset = queryset.filter(
                    device_type__in=form.cleaned_data['device_type'])
            if form.cleaned_data['device_vendor']:
                queryset = queryset.filter(
                    device_vendor__in=form.cleaned_data['device_vendor'])
            return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = DeviceSearchForm(self.request.GET)
        context['banner'] = 'List devices'
        return context


class DeviceUpdateView(UpdateView):
    model = Device
    form_class = DeviceForm
    template_name = "update_device.html"
    success_url = reverse_lazy("list_device")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_success_url(self):
        success_url = self.request.GET.get('next', reverse('list_device'))
        return success_url

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device"
        return context


class DeviceDeleteView(DeleteView):
    model = Device
    template_name = "list_device.html"
    success_url = reverse_lazy("list_device")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceDetailView(DetailView):
    model = Device
    template_name = "detail_device.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.get_object()
        fields = [(field.verbose_name, field.value_from_object(obj))
                  for field in obj._meta.fields]
        context['fields'] = fields
        context['banner'] = 'Device detail informations'
        return context


@login_required()
def create_multiple_device(request):
    banner = "Create device from file"
    try:
        if request.method == "POST" and request.FILES.get("upload-file"):
            uploaded_file = request.FILES["upload-file"]
            if str(uploaded_file).lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(uploaded_file)
                worksheet_01 = wb["Device"]
                worksheet_02 = wb["DeviceManagement"]
                worksheet_03 = wb["DeviceRackLayout"]
                for item in worksheet_01.iter_rows(min_row=2, values_only=True):
                    item = ["" if i is None else i for i in item]
                    DeviceBranch.objects.update_or_create(device_branch=item[2], defaults={
                        'user_created': str(request.user)})
                    DeviceProvince.objects.update_or_create(device_province=item[3], defaults={
                                                            'user_created': str(request.user)})
                    DeviceType.objects.update_or_create(device_type=item[4], defaults={
                                                        'user_created': str(request.user)})
                    DeviceCategory.objects.update_or_create(device_category=item[5], defaults={
                                                            'user_created': str(request.user)})
                    DeviceVendor.objects.update_or_create(device_vendor=item[6], defaults={
                        'user_created': str(request.user)})
                    DeviceOS.objects.update_or_create(device_os=item[7], defaults={
                        'user_created': str(request.user)})
                    DeviceGroup.objects.update_or_create(device_group=item[8], defaults={
                        'user_created': str(request.user)})
                    device_stack = item[9]
                    if device_stack == 'TRUE' or device_stack == 'YES':
                        device_stack = True
                    else:
                        device_stack = False
                    Device.objects.update_or_create(
                        device_ip=item[1],
                        defaults={
                            'device_name': item[0],
                            'device_branch': DeviceBranch.objects.get(device_branch=item[2]),
                            'device_province': DeviceProvince.objects.get(device_province=item[3]),
                            'device_type': DeviceType.objects.get(device_type=item[4]),
                            'device_category': DeviceCategory.objects.get(device_category=item[5]),
                            'device_vendor': DeviceVendor.objects.get(device_vendor=item[6]),
                            'device_os': DeviceOS.objects.get(device_os=item[7]),
                            'device_group': DeviceGroup.objects.get(device_group=item[8]),
                            'device_stack': device_stack,
                            'device_description': item[10],
                            'user_created': str(request.user)
                        }
                    )
                for item in worksheet_02.iter_rows(min_row=2, values_only=True):
                    obj_count_01 = Device.objects.filter(
                        device_ip=item[1]).count()
                    if obj_count_01 > 0:
                        DeviceManagement.objects.update_or_create(
                            device_serial_number=item[2],
                            defaults={
                                'device_ip': Device.objects.get(device_ip=item[1]),
                                'start_ma_date': item[3],
                                'end_ma_date': item[4],
                                'start_license_date': item[5],
                                'end_license_date': item[6],
                                'end_sw_support_date': item[7],
                                'end_hw_support_date': item[8],
                                'start_used_date': item[9],
                                'user_created': str(request.user)
                            }
                        )
                for item in worksheet_03.iter_rows(min_row=2, values_only=True):
                    item = ["" if i is None else i for i in item]
                    obj_count_01 = Device.objects.filter(
                        device_ip=item[1]).count()
                    obj_count_02 = DeviceManagement.objects.filter(
                        device_serial_number=item[2]).count()
                    if obj_count_01 > 0 and obj_count_02 > 0:
                        DeviceRackLayout.objects.update_or_create(
                            device_serial_number=DeviceManagement.objects.get(
                                device_serial_number=item[2]),
                            defaults={
                                'device_ip': Device.objects.get(device_ip=item[1]),
                                'device_rack_name': item[3],
                                'device_rack_unit': item[4],
                                'user_created': str(request.user)
                            }
                        )
                messages.add_message(
                    request, constants.SUCCESS, 'Upload file success')
            else:
                messages.add_message(request, constants.ERROR, f"Only support file type *.xlsx")
    except Exception as error:
        messages.add_message(request, constants.ERROR,
                             f"An error occurred: {error}")
    return render(request, "create_multiple_device.html", context={'banner': banner})


@login_required()
def device_dashboard(request):
    db_title = {
        "db_01": "Count device by OS",
        "db_02": "Count device by vendor",
        "db_04": "Count device by type",
        "db_05": "Count device by province",
        "db_06": "Count device incorrect firmware by type",
        "db_07": "Device management report (Expired after 6 months)",
    }
    return render(request, template_name="device_dashboard.html", context=db_title)


class DeviceManagementListView(ListView):
    model = DeviceManagement
    form_class = DeviceManagementForm
    context_object_name = "devices"
    template_name = "list_device_management.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        form = DeviceManagementSearchForm(self.request.GET)
        if not form.is_valid() or not any(form.cleaned_data.values()):
            return DeviceManagement.objects.none()
        else:
            if form.cleaned_data['device_branch']:
                queryset = queryset.filter(
                    device_ip__device_branch__in=form.cleaned_data['device_branch'])
            if form.cleaned_data['device_ip']:
                queryset = queryset.filter(
                    device_ip=form.cleaned_data['device_ip'])
            return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = DeviceManagementSearchForm(self.request.GET)
        context['banner'] = "List device managements"
        return context


class DeviceManagementCreateView(CreateView):
    model = DeviceManagement
    form_class = DeviceManagementForm
    template_name = "create_device_management.html"
    success_url = reverse_lazy("list_device_management")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device management"
        return context


class DeviceManagementDeleteView(DeleteView):
    model = DeviceManagement
    template_name = "list_device.html"
    success_url = reverse_lazy("list_device_management")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceRackLayoutListView(ListView):
    model = DeviceRackLayout
    form_class = DeviceRackLayoutForm
    context_object_name = "devices"
    template_name = "list_device_rack_layout.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device rack layouts"
        return context


class DeviceRackLayoutCreateView(CreateView):
    model = DeviceRackLayout
    form_class = DeviceRackLayoutForm
    template_name = "create_device_rack_layout.html"
    success_url = reverse_lazy("list_device_rack_layout")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device rack layout"
        return context


class DeviceManagementUpdateView(UpdateView):
    model = DeviceManagement
    form_class = DeviceManagementForm
    template_name = "update_device_management.html"
    success_url = reverse_lazy("list_device_management")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device management"
        return context


class DeviceManagementDetailView(DetailView):
    model = DeviceManagement
    template_name = "detail_device_management.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.get_object()
        fields = [(field.verbose_name, field.value_from_object(obj))
                  for field in obj._meta.fields]
        context['fields'] = fields
        context['banner'] = 'Device management informations'
        return context


class DeviceRackLayoutUpdateView(UpdateView):
    model = DeviceRackLayout
    form_class = DeviceRackLayoutForm
    template_name = "update_device_rack_layout.html"
    success_url = reverse_lazy("list_device_rack_layout")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device rack layout"
        return context


class DeviceRackLayoutDetailView(DetailView):
    model = DeviceRackLayout
    template_name = "detail_device_rack_layout.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.get_object()
        fields = [(field.verbose_name, field.value_from_object(obj))
                  for field in obj._meta.fields]
        context['fields'] = fields
        context['banner'] = 'Device rack layout informations'
        return context


class DeviceRackLayoutDeleteView(DeleteView):
    model = DeviceRackLayout
    template_name = "list_device_rack_layout.html"
    success_url = reverse_lazy("list_device_rack_layout")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceOSCreateView(CreateView):
    model = DeviceOS
    form_class = DeviceOSForm
    template_name = "create_device_os.html"
    success_url = reverse_lazy("list_device_os")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device OS"
        return context


class DeviceOSListView(ListView):
    model = DeviceOS
    context_object_name = "objects"
    template_name = "list_device_os.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device OS"
        return context


class DeviceOSUpdateView(UpdateView):
    model = DeviceOS
    form_class = DeviceOSForm
    template_name = "update_device_os.html"
    success_url = reverse_lazy("list_device_os")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device OS"
        return context


class DeviceOSDeleteView(DeleteView):
    model = DeviceOS
    template_name = "list_device_os.html"
    success_url = reverse_lazy("list_device_os")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


@login_required()
def device_export(request):
    form = DeviceExportForm
    banner = 'Export device'
    data = {"form": form, 'banner': banner}
    try:
        if request.method == "POST":
            form = DeviceExportForm(request.POST)
            if form.is_valid():
                select_id = form.data['database_table']
                if select_id == '1':
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="device.csv"'
                    queryset = Device.objects.all()
                    writer = csv.writer(response)
                    writer.writerow([
                        'device_name',
                        'device_ip',
                        'device_branch',
                        'device_province',
                        'device_type',
                        'device_category',
                        'device_vendor',
                        'device_os',
                        'device_firmware',
                        'device_status',
                        'device_stack',
                        'device_group',
                        'device_description',
                        'device_creation_time',
                        'user_created'
                    ])
                    for item in queryset:
                        writer.writerow([
                            item.device_name,
                            item.device_ip,
                            item.device_branch.device_branch,
                            item.device_province.device_province,
                            item.device_type.device_type,
                            item.device_category.device_category,
                            item.device_vendor.device_vendor,
                            item.device_os.device_os,
                            item.device_firmware,
                            item.device_status,
                            item.device_stack,
                            item.device_group,
                            item.device_description,
                            str(item.device_creation_time),
                            item.user_created
                        ])
                    return response
                elif select_id == '2':
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="device-management.csv"'
                    queryset = DeviceManagement.objects.all()
                    writer = csv.writer(response)
                    writer.writerow([
                        'device_name',
                        'device_ip',
                        'device_serial_number',
                        'start_ma_date',
                        'end_ma_date',
                        'start_license_date',
                        'end_license_date',
                        'end_sw_support_date',
                        'end_hw_support_date',
                        'start_used_date',
                        'user_created'
                    ])
                    for item in queryset:
                        writer.writerow([
                            item.device_ip.device_name,
                            item.device_ip,
                            item.device_serial_number,
                            str(item.start_ma_date),
                            str(item.end_ma_date),
                            str(item.start_license_date),
                            str(item.end_license_date),
                            str(item.end_sw_support_date),
                            str(item.end_hw_support_date),
                            str(item.start_used_date),
                            item.user_created
                        ])
                    return response
                elif select_id == '3':
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="device-rack-layout.csv"'
                    queryset = DeviceRackLayout.objects.all()
                    writer = csv.writer(response)
                    writer.writerow([
                        'device_name',
                        'device_ip',
                        'device_rack_name',
                        'device_rack_unit',
                        'user_created'
                    ])
                    for item in queryset:
                        writer.writerow([
                            item.device_ip.device_name,
                            item.device_ip,
                            item.device_rack_name,
                            item.device_rack_unit,
                            item.user_created
                        ])
                    return response
                elif select_id == '4':
                    list_device_firmware = list(Device.objects.all().values_list(
                        'device_name', 'device_ip', 'device_type__device_type', 'device_firmware'))
                    list_firmware = list(DeviceFirmware.objects.all().values_list(
                        'device_type__device_type', 'firmware'))
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="device-incorrect-firmware.csv"'
                    writer = csv.writer(response)
                    writer.writerow([
                        'device_name',
                        'device_ip',
                        'device_type',
                        'device_firmware'
                    ])
                    for item in list_device_firmware:
                        device_name = item[0]
                        device_ip = item[1]
                        device_type = item[2]
                        device_firmware = item[3]
                        if device_firmware != "":
                            checklist = [i for i in list_firmware if i[0]
                                         == device_type and i[1] == device_firmware]
                            if not checklist:
                                writer.writerow([
                                    device_name,
                                    device_ip,
                                    device_type,
                                    device_firmware
                                ])
                    return response
                elif select_id == '5':
                    datepoint01 = datetime.date.today()
                    datepoint02 = datetime.date.today() + datetime.timedelta(days=180)
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="device-expired-license.csv"'
                    list_device_serial = DeviceManagement.objects.all(
                    ).values_list('device_serial_number', flat=True)
                    writer = csv.writer(response)
                    writer.writerow([
                        'device_branch',
                        'device_province',
                        'device_name',
                        'device_ip',
                        'device_serial',
                        'end_of_date',
                        'description'
                    ])
                    for device_serial in list_device_serial:
                        datalist01 = DeviceManagement.objects.filter(
                            device_serial_number=device_serial, end_ma_date__gte=datepoint01, end_ma_date__lte=datepoint02)
                        datalist02 = DeviceManagement.objects.filter(
                            device_serial_number=device_serial, end_license_date__gte=datepoint01, end_license_date__lte=datepoint02)
                        datalist03 = DeviceManagement.objects.filter(
                            device_serial_number=device_serial, end_sw_support_date__gte=datepoint01, end_sw_support_date__lte=datepoint02)
                        datalist04 = DeviceManagement.objects.filter(
                            device_serial_number=device_serial, end_hw_support_date__gte=datepoint01, end_hw_support_date__lte=datepoint02)
                        if datalist01:
                            for item in datalist01:
                                writer.writerow([
                                    item.device_ip.device_branch,
                                    item.device_ip.device_province,
                                    item.device_ip.device_name,
                                    item.device_ip,
                                    item.device_serial_number,
                                    item.end_ma_date,
                                    'ma_end_of_date'
                                ])
                        if datalist02:
                            for item in datalist02:
                                writer.writerow([
                                    item.device_ip.device_branch,
                                    item.device_ip.device_province,
                                    item.device_ip.device_name,
                                    item.device_ip,
                                    item.device_serial_number,
                                    item.end_license_date,
                                    'license_end_of_date'
                                ])
                        if datalist03:
                            for item in datalist03:
                                writer.writerow([
                                    item.device_ip.device_branch,
                                    item.device_ip.device_province,
                                    item.device_ip.device_name,
                                    item.device_ip,
                                    item.device_serial_number,
                                    item.end_sw_support_date,
                                    'hw_support_end_of_date'
                                ])
                        if datalist04:
                            for item in datalist04:
                                writer.writerow([
                                    item.device_ip.device_branch,
                                    item.device_ip.device_province,
                                    item.device_ip.device_name,
                                    item.device_ip,
                                    item.device_serial_number,
                                    item.end_hw_support_date,
                                    'sw_support_end_of_date'
                                ])
                    return response
                else:
                    messages.add_message(
                        request, constants.ERROR, "Database selected is not valid"
                    )
    except Exception as error:
        messages.add_message(request, constants.ERROR, error)
    return render(request, template_name="device_export.html", context=data)


class DeviceFirmwareCreateView(CreateView):
    model = DeviceFirmware
    form_class = DeviceFirmwareForm
    template_name = "create_device_firmware.html"
    success_url = reverse_lazy("list_device_firmware")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device firmware"
        return context


class DeviceFirmwareListView(ListView):
    model = DeviceFirmware
    context_object_name = "objects"
    template_name = "list_device_firmware.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device firmwares"
        return context


class DeviceFirmwareUpdateView(UpdateView):
    model = DeviceFirmware
    form_class = DeviceFirmwareForm
    template_name = "update_device_firmware.html"
    success_url = reverse_lazy("list_device_firmware")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device firmware"
        return context


class DeviceFirmwareDeleteView(DeleteView):
    model = DeviceFirmware
    template_name = "list_device_firmware.html"
    success_url = reverse_lazy("list_device_firmware")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceBranchCreateView(CreateView):
    model = DeviceBranch
    form_class = DeviceBranchForm
    template_name = "create_device_branch.html"
    success_url = reverse_lazy("list_device_branch")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device branch"
        return context


class DeviceBranchUpdateView(UpdateView):
    model = DeviceBranch
    form_class = DeviceBranchForm
    template_name = "update_device_branch.html"
    success_url = reverse_lazy("list_device_branch")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device branch"
        return context


class DeviceBranchDeleteView(DeleteView):
    model = DeviceBranch
    template_name = "list_device_branch.html"
    success_url = reverse_lazy("list_device_branch")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceBranchListView(ListView):
    model = DeviceBranch
    context_object_name = "objects"
    template_name = "list_device_branch.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device branchs"
        return context


class DeviceGroupCreateView(CreateView):
    model = DeviceGroup
    form_class = DeviceGroupForm
    template_name = "create_device_group.html"
    success_url = reverse_lazy("list_device_group")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Create success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Create device tag"
        return context


class DeviceGroupUpdateView(UpdateView):
    model = DeviceGroup
    form_class = DeviceGroupForm
    template_name = "update_device_group.html"
    success_url = reverse_lazy("list_device_group")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, "Update success")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "Update device tag"
        return context


class DeviceGroupDeleteView(DeleteView):
    model = DeviceGroup
    template_name = "list_device_group.html"
    success_url = reverse_lazy("list_device_group")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(
                self.request, constants.SUCCESS, "Delete success")
        except ProtectedError:
            messages.add_message(
                self.request, constants.ERROR, "This object has been protected"
            )
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)


class DeviceGroupListView(ListView):
    model = DeviceGroup
    context_object_name = "objects"
    template_name = "list_device_group.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banner'] = "List device tags"
        return context
